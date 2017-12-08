import logging

import openpyxl as xl

from . import accountsModule
from . import entityModule
from . import transactionModule
from . import utils

""""
This module parses an input .xlsx file to define the entities, transactions and accounts of the network
"""

logging.basicConfig(level=logging.INFO)

# define the header of data blocks as well as the length (in rows) of each block type
ENTITY_TYPE = "entity"
TRANSACTION_TYPE = "transaction"
ACCOUNT_TYPE = "account"
accounting_data_length = {ENTITY_TYPE: 2, TRANSACTION_TYPE: 6, ACCOUNT_TYPE: 5}

# max number of items that can be defined
MAX_COL_NUMBER = 10000


def is_end_of_data(cell):
    return cell.value is None or cell.value == ""


def get_entity_dict_from_data(data):
    if len(data) != accounting_data_length.get(ENTITY_TYPE):
        return None
    return entityModule.get_entity_dict(*data)


def get_account_dict_from_data(data):
    if len(data) != accounting_data_length.get(ACCOUNT_TYPE):
        return None
    return accountsModule.get_account_dict(*data)


def get_transaction_dict_from_data(data):
    if len(data) != accounting_data_length.get(TRANSACTION_TYPE):
        return None
    reference_accounts, initiator_account, destinatary_account, transfer_ratio, minimum_transfer, maximum_transfer \
        = data[:]
    kargs = dict()
    if reference_accounts is not None:
        kargs["reference_accounts"] = set(reference_accounts.split(";"))
    if minimum_transfer is not None or maximum_transfer is not None:
        kargs["transfer_ratio_bounds"] = (minimum_transfer, maximum_transfer)
    if transfer_ratio is None or not utils.is_float(transfer_ratio):
        transfer_ratio = 0.3
        kargs["transfer_ratio_calculation"] = transactionModule.TRANSFER_RATIO_THEIR
    transaction = transactionModule.get_transaction_dict(initiator_account, destinatary_account, transfer_ratio,
                                                         **kargs)
    return transaction


def get_accounting_object_dict_from_data(data, object_type):
    if object_type == ENTITY_TYPE:
        return get_entity_dict_from_data(data)
    elif object_type == ACCOUNT_TYPE:
        return get_account_dict_from_data(data)
    elif object_type == TRANSACTION_TYPE:
        return get_transaction_dict_from_data(data)


def get_account_object_list(accounting_sheet, row, object_type):
    """
    Get a list of object dictionaries. Object are either "entites", "transactions" or "accounts"
    :param accounting_sheet:
    :param row: current row object
    :param object_type: either "entites", "transactions" or "accounts"
    :return:
    """
    initial_row = row[0].row + 1
    final_row = initial_row + accounting_data_length.get(object_type) - 1
    entities_list = []
    for col in accounting_sheet.iter_cols(min_row=initial_row, max_row=final_row, min_col=2, max_col=MAX_COL_NUMBER):
        if is_end_of_data(col[0]):
            break
        data = []
        for cell in col:
            data.append(cell.value)
        new_entity = get_accounting_object_dict_from_data(data, object_type)
        if new_entity is not None:
            entities_list.append(new_entity)
    return entities_list


def read_accounting_book(path_to_file):
    """
    Parse an .xlsx file to define lists of entities, accounts and transaction that will be used to construct the entity graph
    :param path_to_file: path to the .xlsx file
    :return:
    all_entities_list: list of entity dictionaries used in this entity network
    all_accounts_list: list of all_accounts dictionaries used in this entity network
    all_transactions_list: list of all_transactions dictionaries used in this entity network
    """
    all_entities_list = []
    all_accounts_list = []
    all_transactions_list = []
    accounting_book = xl.load_workbook(path_to_file)
    accounting_sheet = accounting_book.get_sheet_by_name("1_DEFINITIONS")
    # read the first 50 rows, begining by the cell A1
    for row in accounting_sheet.iter_rows(min_row=1, max_col=1, max_row=50):
        if row[0].value == "entities":
            all_entities_list.extend(get_account_object_list(accounting_sheet, row, ENTITY_TYPE))
        elif row[0].value == "accounts":
            all_accounts_list.extend(get_account_object_list(accounting_sheet, row, ACCOUNT_TYPE))
        elif row[0].value == "transactions":
            all_transactions_list.extend(get_account_object_list(accounting_sheet, row, TRANSACTION_TYPE))
    return all_entities_list, all_accounts_list, all_transactions_list
